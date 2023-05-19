import openpyxl
import random
import copy
import math
import itertools
import curses
import os
from openpyxl.styles import Alignment, Font


# this class takes in a text file and extracts the swimmer information from it
class InputProcessor:
    def __init__(self, file_path):
        self.file = openpyxl.load_workbook(file_path, data_only=False)["Sheet1"]
        self.timings = self.get_timings()
        self.num_teams = self.get_num_teams()

    def get_timings(self):
        dummy_timings = []

        for row in self.file.iter_rows(2, self.file.max_row):
            runner_details = [row[1].value]
            for i in range(2, 6):
                runner_details.append(self.get_seconds(row[i].value))
            dummy_timings.append(runner_details)
        return dummy_timings

    def get_num_teams(self):
        division_size = []
        index = 2
        while self.file.cell(row=index, column=9).value is not None:
            division_size.append(self.file.cell(row=index, column=9).value)
            index += 1
        return division_size

    def get_seconds(self, ms):
        if ms == "inf":
            return float("inf")
        if type(ms) is not float and type(ms) is not int:
            time = str(ms)
            if ':' in time:
                split = time.split(":")
                return int(split[len(split) - 2]) * 60 + float(split[len(split) - 1])
                #return int(split[0]) * 60 + float(split[1])
        return ms


def generate_initial_division(race_information, division_no):
    team_index = [0 for i in range(race_information.num_teams[division_no])]
    team_timings = [0 for i in range(race_information.num_teams[division_no])]
    division = [-1 for i in range(race_information.num_teams[division_no] * 4)]

    beginning_index = sum([x for x in race_information.num_teams[0: division_no]]) * 4
    swimmers = copy.deepcopy(race_information.timings[beginning_index : beginning_index + race_information.num_teams[division_no] * 4])

    race_order = [3, 4, 1, 2]
    for race_index in race_order:
        chosen_swimmers = []
        for i in range(race_information.num_teams[division_no]):
            chosen_race_timings = [swimmer[race_index] for swimmer in swimmers]
            chosen_swimmer = swimmers[chosen_race_timings.index(min(chosen_race_timings))]
            swimmers.remove(chosen_swimmer)
            chosen_swimmers.append(chosen_swimmer)

        while len(chosen_swimmers) > 0:
            teams_with_less_swimmers = []
            min_num_swimmers = min(team_index)

            for j in range(len(team_index)):
                if team_index[j] == min_num_swimmers:
                    teams_with_less_swimmers.append(j)

            max_timing = max([team_timings[x] for x in teams_with_less_swimmers])
            slow_teams = []

            for x in teams_with_less_swimmers:
                if team_timings[x] == max_timing:
                    slow_teams.append(x)

            chosen_team = random.choice(slow_teams)

            race_timings = [swimmer[race_index] for swimmer in chosen_swimmers]
            fastest_swimmer = chosen_swimmers[race_timings.index(min(race_timings))]
            chosen_swimmers.remove(fastest_swimmer)

            division[chosen_team * 4 + race_index - 1] = fastest_swimmer
            team_timings[chosen_team] += fastest_swimmer[race_index]
            team_index[chosen_team] += 1

    return division

def get_timings(division, race_information, division_no):
    team_timings = [sum([division[i * 4 + j][j + 1] for j in range(4)]) for i in range(race_information.num_teams[division_no])]
    return team_timings

def get_team_timing(team):
    return sum([team[i][i+1] for i in range(4)])

def get_optimum_timings(division, race_information, division_no):
    team_timings = []
    for i in range(race_information.num_teams[division_no]):
        team = division[i * 4: i * 4 + 4]

        best_timing = float('inf')
        for perm in (itertools.permutations(team)):
            if get_team_timing(perm) < best_timing:
                best_timing = get_team_timing(perm)

        team_timings.append(best_timing)

    return team_timings

def get_optimum_division(division, race_information, division_no):
    best_division = []
    for i in range(race_information.num_teams[division_no]):
        team = division[i * 4: i * 4 + 4]

        best_timing = float('inf')
        for perm in (itertools.permutations(team)):
            if get_team_timing(perm) < best_timing:
                best_team = perm
                best_timing = get_team_timing(perm)

        best_division += best_team

    return best_division

def get_range(division, race_information, division_no):
    team_timing = get_optimum_timings(division, race_information, division_no)
    return max(team_timing) - min(team_timing)

def fitness(division, race_information, division_no):
    range = get_range(division, race_information, division_no)
    if range == 0:
        return float('inf')
    return 1 / range

def create_neighbours(division):
    neighbours = []
    for i in range(len(division)):
        for j in range(i + 1, len(division)):
            new_neighbour = copy.deepcopy(division)
            new_neighbour[i], new_neighbour[j] = new_neighbour[j], new_neighbour[i]
            neighbours.append(new_neighbour)
    return neighbours

def main():
    screen = curses.initscr()
    screen.addstr(0, 0, "This is a program which creates a set of competitive relay teams.")
    start_y, start_x = screen.getyx()
    start_y += 1

    screen.addstr(
        "\nWould you like to manually enter the name of the input file, or choose from the list of excel file in the same directory[1/2]: ")
    end_y, end_x = screen.getyx()
    answer = ""
    while True:
        answer = screen.getstr().decode('utf-8')
        if answer in ['1', '2']:
            break
        screen.move(start_y, 0)
        screen.clrtoeol()
        screen.move(end_y, 0)
        screen.clrtoeol()
        screen.move(start_y, 0)
        screen.addstr(
            "Invalid please choose again. Would you like to manually enter the name of the input file, or choose from the list of excel file in the same directory[1/2]: ")

    valid_endings = ['xls', 'xlsx']
    excel_files = [file for file in os.listdir(os.getcwd()) if file.endswith(tuple(valid_endings))]

    if answer == '1':
        screen.clear()
        screen.addstr(0, 0, "Please enter the file containing the runners' names, including the file type: ")
        start_y, start_x = 0, 0
        end_y, end_x = screen.getyx()

        while True:
            path = screen.getstr().decode('utf-8')

            if path in excel_files:
                break

            screen.move(start_y, 0)
            screen.clrtoeol()
            screen.move(end_y, 0)
            screen.clrtoeol()
            screen.move(start_y, 0)
            screen.addstr(
                "File not found in the directory. Please try again and enter the file name, including the file type: ")
    else:
        screen.clear()
        screen.addstr(0, 0,
                      "Here are the valid files for input in the directory. To choose a file, type its corresponding number: ")

        for index, file in enumerate(excel_files):
            screen.addstr(index + 1, 5, str(index + 1) + ") " + file)
        start_y, start_x = screen.getyx()
        start_y += 2
        screen.addstr("\n\nFile choice: ")
        end_y, end_x = screen.getyx()
        while True:
            file_no = screen.getstr().decode('utf-8')

            if file_no.isnumeric():
                if len(excel_files) >= eval(file_no) > 0:
                    break

            screen.move(start_y, 0)
            screen.clrtoeol()
            screen.move(end_y, 0)
            screen.clrtoeol()
            screen.move(start_y, 0)
            screen.addstr("Invalid input. To choose a file, type its corresponding number: ")

        path = excel_files[eval(file_no) - 1]

    information = InputProcessor(str(path))
    reset_time = 2
    num_competitive_runners = sum([no_of_teams for no_of_teams in information.num_teams]) * 4

    final_divisions = []
    if num_competitive_runners > len(information.timings):
        screen.clear()
        screen.addstr(0, 0, 'Insufficient number of runners. Press ENTER to quit')
        screen.refresh()
        screen.getch()
        curses.endwin()
        return
    elif num_competitive_runners < len(information.timings):
        screen.clear()
        screen.addstr(0, 0, "Too many runners. The following runners will not be included in the team making process: ")
        for x in information.timings[num_competitive_runners::]:
            screen.addstr('\n\t' + x[0])
        screen.addstr("\nDo you want to continue with this elimination? [y to continue/anykey to quit] : ")
        answer = screen.getstr().decode('utf-8')
        if answer != 'y':
            curses.endwin()
            return
    screen.clear()
    screen.addstr(0, 0, "CREATING TEAMS")

    for j in range(len(information.num_teams)):
        initial_division = current_division = generate_initial_division(information, j)
        num_iteration = 1000
        initial_temperature = current_temperature = 10
        cooling_rate = 0.99
        limit = 5
        for i in range(num_iteration):
            if (i / num_iteration) * 100 >= limit:
                screen.move(j + 1, 0)
                screen.clrtoeol()
                screen.addstr(j + 1, 0,
                              str(int((i * 100) / num_iteration)) + "% through to completing teams for division " + str(
                                  j + 1) + " " + "[" + "|" * int((i * 20) / num_iteration) + " " * math.ceil(
                                  20 - (i * 20) / num_iteration) + "]")
                screen.refresh()
                limit += 5

            current_temperature *= cooling_rate

            neighbours = create_neighbours(current_division)
            fitness_scores = [fitness(neighbour, information, j) for neighbour in neighbours]
            chosen_division = neighbours[fitness_scores.index(max(fitness_scores))]
            chosen_fitness = fitness(chosen_division, information, j)
            current_fitness = fitness(current_division, information, j)

            if chosen_fitness >= current_fitness:
                current_division = chosen_division
            else:
                probability = 1 / (1 - math.exp((chosen_fitness - current_fitness) / current_temperature))
                if random.random() < probability:
                    current_division = chosen_division

        screen.move(j + 1, 0)
        screen.clrtoeol()
        screen.addstr(j + 1, 0,
                      "100% through to completing teams for division " + str(j + 1) + " " + "[" + "|" * 20 + "]")
        screen.refresh()

        current_division = get_optimum_division(current_division, information, j)
        final_divisions.append(current_division)

    screen.clear()
    screen.addstr(0, 0, 'OUTPUT')
    screen.addstr(1, 0,
                  "Please enter the name of the excel file to create and output the teams into including the file type. For default name, press enter instead: ")
    start_y, start_x = 1, 0
    end_y, end_x = screen.getyx()
    while True:
        answer = screen.getstr().decode()

        if answer.endswith(tuple(valid_endings)) or answer == '':
            break

        screen.move(start_y, 0)
        screen.clrtoeol()
        screen.move(end_y, 0)
        screen.clrtoeol()
        screen.move(start_y, 0)
        screen.addstr(
            "Invalid file type. Please enter the name of the excel file to create and output the teams into including the file type. For default name, press enter instead: ")

    wb = openpyxl.Workbook()
    sheet = wb.active

    if answer == '':
        answer = "teams.xlsx"

    sheet.merge_cells('A1:P1')
    sheet.cell(row = 1, column=1).value = 'J.S.A.S.A'
    sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal='center')
    sheet.cell(row = 1, column = 1).font = Font(name = "Courier New", size=10, bold=True)

    sheet.merge_cells('A2:P2')
    sheet.cell(row = 2, column = 1).value = 'Department of Physical Education'
    sheet.cell(row = 2, column = 1).alignment = Alignment(horizontal = 'center')
    sheet.cell(row = 2, column = 1).font = Font(name = "Courier New", size=10, bold=True)

    sheet.merge_cells('A3:P3')
    sheet.cell(row = 3, column = 1).value = "MEN'S AQUATICS -- "
    sheet.cell(row = 3, column = 1).alignment = Alignment(horizontal = 'center')
    sheet.cell(row = 3, column = 1).font = Font(name = "Courier New", size=14, bold=True)

    sheet.merge_cells('A5:P5')
    sheet.cell(row=5, column=1).value = "TEAMS FOR MEDLEY RELAY"
    sheet.cell(row = 5, column = 1).alignment = Alignment(horizontal = 'center')
    sheet.cell(row = 5, column = 1).font = Font(name = "Calibri", size=14, bold=True)




    races = ["Back", "Breast", "B-fly", "Crawl"]
    lane_order = [3,4,2,5,1,6]
    section_len = 10
    header_offset = 6
    for index, division in enumerate(final_divisions):
        sheet.cell(row=index * section_len + 1 + header_offset, column=1).value = "Division " + str(index + 1)
        sheet.cell(row=index * section_len + 1 + header_offset, column=1).font = Font(bold = True)
        current_lanes = lane_order[0:information.num_teams[index]]
        current_lanes.sort()
        for i in range(information.num_teams[index]):
            sheet.cell(row=index * section_len + 2 + header_offset, column=i* 3 + 3).value = "LANE" + str(current_lanes[i])

        for k in range(5):
            if k < 4:
                sheet.cell(row = index * section_len + k + 4 + header_offset, column = 1).value = races[k]
            for l in range(information.num_teams[index]):
                if k < 4:
                    sheet.cell(row=index * section_len + k + 4 + header_offset, column=l * 3 + 3).value = division[l * 4 + k][0]
                    sheet.cell(row=index * section_len + k + 4 + header_offset, column=l * 3 + 4).value = division[l * 4 + k][k + 1]
                else:
                    sheet.cell(row=index * section_len + k + 4 + header_offset, column=l * 3 + 4).value = get_timings(division, information, index)[l]
            sheet.cell(row = index * section_len + 5 + 4 + header_offset, column=1).value = 'RESULTS:'
            sheet.cell(row=index * section_len + 5 + 4 + header_offset, column=1).font = Font(bold = True)

    wb.save(answer)

    screen.clear()
    screen.addstr(0, 0, "The teams are created and outputted to " + answer + ". Press enter to quit.")
    screen.getch()
    curses.endwin()


main()
